import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def merge_with_existing2(df_new, excel_path, date):
    wynik_col = f"Wynik {date}"

    if os.path.exists(excel_path):
        df_existing = pd.read_excel(excel_path)

        for col in ["Min", "Max"]:
            if col in df_existing.columns:
                df_existing[col] = pd.to_numeric(df_existing[col], errors="coerce")
            if col in df_new.columns:
                df_new[col] = pd.to_numeric(df_new[col], errors="coerce")

        if wynik_col not in df_existing.columns:
            df_existing[wynik_col] = None

        for _, row in df_new.iterrows():
            mask = (
                (df_existing["Badanie"] == row["Badanie"]) &
                (df_existing["Jedn."] == row["Jedn."]) &
                (df_existing["Zakres referencyjny"] == row["Zakres referencyjny"])
            )
            if mask.any():
                df_existing.loc[mask, wynik_col] = row[wynik_col]
            else:
                new_row = {
                    "Badanie": row["Badanie"],
                    "Jedn.": row["Jedn."],
                    "Zakres referencyjny": row["Zakres referencyjny"],
                    "Min": row.get("Min", np.nan),
                    "Max": row.get("Max", np.nan),
                    wynik_col: row[wynik_col]
                }
                df_existing = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)

        df_merged = df_existing
    else:
        df_merged = df_new

    return df_merged


def apply_color_formatting(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    wynik_cols = [i + 1 for i, h in enumerate(headers) if h.startswith("Wynik ")]
    min_idx = headers.index("Min") + 1 if "Min" in headers else None
    max_idx = headers.index("Max") + 1 if "Max" in headers else None

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        try:
            min_val = float(str(row[min_idx-1].value).replace(",", ".")) if min_idx and row[min_idx-1].value not in [None, ""] else None
        except:
            min_val = None
        try:
            max_val = float(str(row[max_idx-1].value).replace(",", ".")) if max_idx and row[max_idx-1].value not in [None, ""] else None
        except:
            max_val = None

        for col_idx in wynik_cols:
            cell = row[col_idx-1]
            try:
                val = float(str(cell.value).replace(",", "."))
            except:
                continue
            if (min_val is not None and val < min_val) or (max_val is not None and val > max_val):
                cell.fill = red_fill
            else:
                cell.fill = green_fill

    wb.save(excel_path)
